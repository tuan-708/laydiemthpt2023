from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook

listSubject = []

def getScore(sbd):
    soup = BeautifulSoup(urlopen(f"https://vietnamnet.vn/giao-duc/diem-thi/tra-cuu-diem-thi-tot-nghiep-thpt/2023/{sbd}.html"), "html.parser")
    mydivs = soup.find_all("td")
    data = {"sbd": sbd}
    data_dict = []
    subject = ""
    for i,v in enumerate(mydivs):
        if i%2 == 0:
            subject = mydivs[i].text
            if subject not in listSubject:
                listSubject.append(subject)
        else:
            score = mydivs[i].text
            core = {subject: score}
            data_dict.append(core)
    data["score"] = data_dict
    print(data)
    return  data

studentsCore = []

for i in range(10000001, 100000000):
    try:
        studentsCore.append(getScore(i))
    except:
        break

file_source = 'score.xlsx'
workbook = load_workbook(filename=file_source)
ws4 = workbook["Sheet1"]
ws4.cell(row = 1, column = 1).value = 'Số báo danh'
ws4.cell(row = 1, column = 2).value = 'Toán'
ws4.cell(row = 1, column = 3).value = 'Lí'
ws4.cell(row = 1, column = 4).value = 'Hóa'
ws4.cell(row = 1, column = 5).value = 'Sinh'
ws4.cell(row = 1, column = 6).value = 'Văn'
ws4.cell(row = 1, column = 7).value = 'Ngoại ngữ'
ws4.cell(row = 1, column = 8).value = 'Sử'
ws4.cell(row = 1, column = 9).value = 'Địa'
ws4.cell(row = 1, column = 10).value = 'GDCD'

for i, v in enumerate(studentsCore):
    ws4.cell(row = i+2, column = 1).value = v['sbd']
    for s in v['score']:
        if s.get('Toán') is not None:
            ws4.cell(row=i+2, column=2).value = s.get('Toán')
        if s.get('Lí') is not None:
            ws4.cell(row=i+2, column=3).value = s.get('Lí')
        if s.get('Hóa') is not None:
            ws4.cell(row=i+2, column=4).value = s.get('Hóa')
        if s.get('Sinh') is not None:
            ws4.cell(row=i+2, column=5).value = s.get('Sinh')
        if s.get('Văn') is not None:
            ws4.cell(row=i+2, column=6).value = s.get('Văn')
        if s.get('Ngoại ngữ') is not None:
            ws4.cell(row=i+2, column=7).value = s.get('Ngoại ngữ')
        if s.get('Sử') is not None:
            ws4.cell(row=i+2, column=8).value = s.get('Sử')
        if s.get('Địa') is not None:
            ws4.cell(row=i+2, column=9).value = s.get('Địa')
        if s.get('GDCD') is not None:
            ws4.cell(row=i+2, column=10).value = s.get('GDCD')

workbook.save(filename=file_source)